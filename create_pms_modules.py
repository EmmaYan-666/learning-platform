#!/usr/bin/env python3
"""
酒店PMS系统模块规划 - Excel生成器
"""

import subprocess
import sys

def check_and_install_dependencies():
    """检查并安装所需的Python库"""
    print("检查依赖库...")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        print("✓ openpyxl已安装")
        return True
    except ImportError:
        print("× openpyxl未安装,正在安装...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            print("✓ openpyxl安装成功")
            return True
        except subprocess.CalledProcessError as e:
            print(f"✗ 安装失败: {e}")
            print("\n请手动安装: pip3 install openpyxl")
            return False

def create_excel():
    """创建酒店PMS系统模块规划Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "PMS系统模块规划"
    
    # 定义样式
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E2761', end_color='1E2761', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    level1_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    level1_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    
    level2_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    level2_font = Font(name='Arial', size=10, bold=False, color='000000')
    
    normal_font = Font(name='Arial', size=10)
    normal_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 30
    
    # 表头
    headers = ['一级导航', '二级导航', '功能说明', '主要功能点']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    # PMS系统模块数据
    pms_modules = [
        # 前台管理
        ('前台管理', '预订管理', '处理客房预订、修改、取消等操作', '新建预订、修改预订、取消预订、预订查询、预订确认、超售管理'),
        ('前台管理', '入住登记', '办理客人入住手续', '散客入住、团队入住、VIP入住、快速入住、预订单入住'),
        ('前台管理', '房态管理', '实时查看和管理客房状态', '房态图、房态统计、客房分配、换房操作、房间锁定/解锁'),
        ('前台管理', '退房结账', '办理客人退房和结算', '快速退房、账单结算、退款处理、发票打印、离店登记'),
        ('前台管理', '前台收银', '处理日常收银业务', '收款、退款、预收款管理、押金管理、日结报表'),
        
        # 客房管理
        ('客房管理', '房务中心', '客房服务调度和管理', '客房服务请求、维修报修、失物招领、客房检查'),
        ('客房管理', '清洁管理', '客房清洁排班和记录', '清洁排班、清洁记录、房间检查、清洁质量评分'),
        ('客房管理', '物品管理', '客房用品和布草管理', '布草管理、客用品管理、借用品管理、损耗统计'),
        ('客房管理', '维修管理', '客房设施维修管理', '维修报修、维修派工、维修记录、维修统计'),
        
        # 销售管理
        ('销售管理', '客户管理', '客户信息和关系维护', '客户档案、客户分级、客户标签、客户跟进记录'),
        ('销售管理', '协议公司', '协议客户管理', '协议公司档案、协议价格、协议合同、协议用量统计'),
        ('销售管理', '销售机会', '销售线索和商机管理', '销售线索、销售机会、跟进记录、销售漏斗'),
        ('销售管理', '团队预订', '团队客户预订管理', '团队预订、团队排房、团队账务、团队确认'),
        
        # 财务管理
        ('财务管理', '账务管理', '日常账务处理', '账务录入、账务调整、账务查询、账务审核'),
        ('财务管理', '应收管理', '应收账款管理', '应收账单、催款管理、账龄分析、坏账处理'),
        ('财务管理', '财务报表', '财务相关报表', '收入报表、营收分析、现金流量表、资产负债表'),
        ('财务管理', '夜审管理', '夜间审核流程', '夜审准备、夜审执行、夜审报表、日结处理'),
        
        # 会员管理
        ('会员管理', '会员档案', '会员信息管理', '会员注册、会员信息、会员等级、会员积分'),
        ('会员管理', '积分管理', '会员积分管理', '积分累计、积分兑换、积分调整、积分明细'),
        ('会员管理', '权益管理', '会员权益配置', '权益配置、权益发放、权益使用、权益统计'),
        ('会员管理', '会员营销', '会员营销活动', '营销活动、优惠券管理、短信营销、会员关怀'),
        
        # 报表中心
        ('报表中心', '运营报表', '日常运营数据报表', '入住率报表、RevPAR分析、客源分析、房态报表'),
        ('报表中心', '财务报表', '财务相关报表', '收入日报、收入月报、AR账龄表、现金流报表'),
        ('报表中心', '销售报表', '销售业绩报表', '销售业绩、客户分析、渠道分析、市场分析'),
        ('报表中心', '客房报表', '客房运营报表', '清洁报表、维修报表、布草报表、物品消耗报表'),
        
        # 系统设置
        ('系统设置', '基础配置', '系统基础参数设置', '酒店信息、部门设置、员工管理、班次设置'),
        ('系统设置', '房型管理', '房型配置管理', '房型定义、房号管理、房价设置、房态配置'),
        ('系统设置', '价格管理', '房价体系管理', '基础房价、价格策略、促销价格、协议价格'),
        ('系统设置', '权限管理', '用户权限配置', '角色管理、权限分配、用户管理、操作日志'),
        
        # 接口管理
        ('接口管理', '渠道管理', 'OTA渠道对接管理', '携程、美团、飞猪、Booking、Agoda等渠道配置'),
        ('接口管理', '门锁接口', '门锁系统对接', '门锁品牌配置、发卡接口、门锁日志'),
        ('接口管理', '支付接口', '支付系统对接', '微信支付、支付宝、银联、预授权接口'),
        ('接口管理', 'PMS接口', '第三方系统对接', '中央预订系统(CRS)、会员系统、财务系统'),
        
        # 移动端
        ('移动端', '自助入住', '客人自助服务', '自助入住机、手机入住、自助选房、自助退房'),
        ('移动端', '微信小程序', '微信端服务', '微信预订、微信入住、微信服务、微信支付'),
        ('移动端', '员工APP', '移动办公', '房态查询、服务派单、移动审批、消息通知'),
    ]
    
    # 写入数据
    current_row = 2
    current_level1 = None
    level1_start_row = None
    
    for idx, (level1, level2, desc, features) in enumerate(pms_modules):
        row = current_row
        
        # 一级导航
        if level1 != current_level1:
            if current_level1 is not None and level1_start_row is not None:
                # 合并上一个一级导航单元格
                if current_row - 1 > level1_start_row:
                    ws.merge_cells(start_row=level1_start_row, start_column=1, end_row=current_row-1, end_column=1)
            
            current_level1 = level1
            level1_start_row = row
            
            cell = ws.cell(row=row, column=1, value=level1)
            cell.font = level1_font
            cell.fill = level1_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        else:
            cell = ws.cell(row=row, column=1, value='')
            cell.border = thin_border
        
        # 二级导航
        cell = ws.cell(row=row, column=2, value=level2)
        cell.font = level2_font
        cell.fill = level2_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # 功能说明
        cell = ws.cell(row=row, column=3, value=desc)
        cell.font = normal_font
        cell.alignment = normal_alignment
        cell.border = thin_border
        
        # 主要功能点
        cell = ws.cell(row=row, column=4, value=features)
        cell.font = normal_font
        cell.alignment = normal_alignment
        cell.border = thin_border
        
        ws.row_dimensions[row].height = 35
        current_row += 1
    
    # 合并最后一个一级导航单元格
    if level1_start_row is not None and current_row - 1 > level1_start_row:
        ws.merge_cells(start_row=level1_start_row, start_column=1, end_row=current_row-1, end_column=1)
    
    # 创建第二个工作表 - 模块分类汇总
    ws2 = wb.create_sheet(title="模块分类汇总")
    
    # 设置列宽
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 60
    
    # 表头
    summary_headers = ['一级模块', '二级模块数量', '主要功能']
    for col, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws2.row_dimensions[1].height = 30
    
    # 汇总数据
    summary_data = [
        ('前台管理', 5, '预订、入住、房态、退房、收银'),
        ('客房管理', 4, '房务中心、清洁、物品、维修'),
        ('销售管理', 4, '客户、协议公司、销售机会、团队预订'),
        ('财务管理', 4, '账务、应收、报表、夜审'),
        ('会员管理', 4, '档案、积分、权益、营销'),
        ('报表中心', 4, '运营、财务、销售、客房报表'),
        ('系统设置', 4, '基础配置、房型、价格、权限'),
        ('接口管理', 4, '渠道、门锁、支付、PMS接口'),
        ('移动端', 3, '自助入住、微信小程序、员工APP'),
    ]
    
    for idx, (module, count, functions) in enumerate(summary_data, 2):
        cell = ws2.cell(row=idx, column=1, value=module)
        cell.font = level1_font
        cell.fill = level1_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        
        cell = ws2.cell(row=idx, column=2, value=count)
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        cell = ws2.cell(row=idx, column=3, value=functions)
        cell.font = normal_font
        cell.alignment = normal_alignment
        cell.border = thin_border
        
        ws2.row_dimensions[idx].height = 30
    
    # 添加总计行
    total_row = len(summary_data) + 2
    cell = ws2.cell(row=total_row, column=1, value='总计')
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.alignment = center_alignment
    cell.border = thin_border
    
    cell = ws2.cell(row=total_row, column=2, value=36)
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.alignment = center_alignment
    cell.border = thin_border
    
    cell = ws2.cell(row=total_row, column=3, value='9个一级模块，36个二级模块')
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.alignment = normal_alignment
    cell.border = thin_border
    
    # 保存文件
    output_path = "/Users/emma/WorkBuddy/20260312092945/酒店PMS系统模块规划.xlsx"
    wb.save(output_path)
    print(f"\n✓ Excel文件已创建: {output_path}")
    return output_path

def main():
    """主函数"""
    print("=" * 60)
    print("酒店PMS系统模块规划 - Excel生成器")
    print("=" * 60)
    print()
    
    # 检查并安装依赖
    if not check_and_install_dependencies():
        sys.exit(1)
    
    # 创建Excel
    output_file = create_excel()
    
    print("\n" + "=" * 60)
    print("完成! 包含9个一级模块,36个二级模块")
    print("=" * 60)

if __name__ == "__main__":
    main()
